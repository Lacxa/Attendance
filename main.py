import json
import time
from datetime import datetime
import os
from threading import Thread

from kivy import utils
from kivy import platform
from kivy.uix.floatlayout import FloatLayout
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from kivy.base import EventLoop
from kivy.clock import Clock, mainthread
from kivy.core.window import Window
from kivy.properties import NumericProperty, StringProperty, ObjectProperty
from kivymd.app import MDApp
from kivymd.toast import toast
from kivymd.uix.pickers import MDDatePicker
from openpyxl.styles.builtins import total

from database import Uzuri as Uz

import kivymd_extensions.akivymd


Window.keyboard_anim_args = {"d": .2, "t": "linear"}
Window.softinput_mode = "below_target"
Clock.max_iteration = 250


if utils.platform != 'android':
    Window.size = [420, 740]


class MainApp(MDApp):
    # APP
    screens = ['home']
    screens_size = NumericProperty(len(screens) - 1)
    current = StringProperty(screens[len(screens) - 1])
    selected_date = StringProperty("")
    report = StringProperty("")
    filepath = StringProperty("")

    internet = StringProperty("asset/slide_three_img.png")
    nodata = StringProperty("asset/slide_three_img.png")

    #Database
    parent_data = Uz.parent(Uz())
    visitor_data = Uz.guest(Uz())
    school_data = Uz.school(Uz())
    saved = Uz.fetch_all_parent(Uz())

    selected_parent = []
    selected_visitor = []
    selected_school = []

    @mainthread
    def on_start(self):
        if utils.platform == 'android':
            self.request_android_permissions()

        self.keyboard_hooker()
        #self.theme_cls.theme_style = "Dark"

        self.cal()
        self.display()
        self.month()

    @mainthread
    def refresh(self):
        self.parent_data = Uz.parent(Uz())
        self.visitor_data = Uz.guest(Uz())
        self.school_data = Uz.school(Uz())

        self.cal()
        self.display()

    """ KEYBOARD INTEGRATION """

    def keyboard_hooker(self, *args):
        EventLoop.window.bind(on_keyboard=self.hook_keyboard)

    def hook_keyboard(self, window, key, *largs):
        print(self.screens_size)
        if key == 27 and self.screens_size > 0:
            print(f"your were in {self.current}")
            last_screens = self.current
            self.screens.remove(last_screens)
            print(self.screens)
            self.screens_size = len(self.screens) - 1
            self.current = self.screens[len(self.screens) - 1]
            self.screen_capture(self.current)
            return True
        elif key == 27 and self.screens_size == 0:
            toast('Press Home button!')
            return True

    """ SCREEN FUNCTIONS """

    def screen_capture(self, screen):
        sm = self.root
        sm.current = screen
        if screen in self.screens:
            pass
        else:
            self.screens.append(screen)
        print(self.screens)
        self.screens_size = len(self.screens) - 1
        self.current = self.screens[len(self.screens) - 1]
        print(f'size {self.screens_size}')
        print(f'current screen {screen}')

    def screen_leave(self):
        print(f"your were in {self.current}")
        last_screens = self.current
        self.screens.remove(last_screens)
        print(self.screens)
        self.screens_size = len(self.screens) - 1
        self.current = self.screens[len(self.screens) - 1]
        self.screen_capture(self.current)

    """ DISPENSING FUNCTIONS """

    def update_graph(self):
        chart3 = self.root.ids.chart3
        chart3.x_labels =  ["Visitors", "Schools", "Students"]
        chart3.y_labels = ["45", "50", "65"]
        chart3.update()

    def display(self):
        self.root.ids.prt.data = {}
        self.root.ids.vis.data = {}
        self.root.ids.sch.data = {}

        self.root.ids.pvs.data = {}

        self.root.ids.pvs.data.append(
            {
                "viewclass": "InfoCard",
                "name": "Parents",
            }
        )
        self.root.ids.pvs.data.append(
            {
                "viewclass": "InfoCard",
                "name": "Visitors",
            }
        )
        self.root.ids.pvs.data.append(
            {
                "viewclass": "InfoCard",
                "name": "Schools",
            }
        )
        if self.parent_data:
            self.root.ids.prt.data = {}
            for i, y in self.parent_data.items():
                for child in y["children"]:
                    self.root.ids.prt.data.append(
                        {
                            "viewclass": "StuCard",
                            "name": child["name"],
                            "pname": y["name"],
                            "pnumber": y["phone"],
                            "cage": child["age"],
                        }
                    )

        else:
            self.root.ids.prt.data.append(
                {
                    "viewclass": "StuCardx",
                    "name": "No data available yet!",
                }
            )

        if self.visitor_data:
            self.root.ids.vis.data = {}
            for i, y  in self.visitor_data.items():
                self.root.ids.vis.data.append(
                    {
                        "viewclass": "VisCard",
                        "name": y["company_name"],
                        "gname": y["guider_name"],
                        "total": y["number_of_visitors"],
                        "phone": y["phone_number"],
                    }
                )

        else:
            self.root.ids.vis.data.append(
                {
                    "viewclass": "StuCardx",
                    "name": "No data available yet!",
                }
            )

        if self.school_data:
            self.root.ids.sch.data = {}
            for i, y  in self.school_data.items():
                self.root.ids.sch.data.append(
                    {
                        "viewclass": "SchCard",
                        "name": y["school_name"],
                        "gname": y["teacher_name"],
                        "total": y["number_of_students"],
                        "phone": y["phone_number"],
                    }
                )

        else:
            self.root.ids.sch.data.append(
                {
                    "viewclass": "StuCardx",
                    "name": "No data available yet!",
                }
            )

        self.stop()


    vis = StringProperty()
    par = StringProperty()
    sch = StringProperty()
    count = StringProperty()

    def cal(self):
        self.selected_parent = self.parent_data
        self.selected_visitor = self.visitor_data
        self.selected_school = self.school_data

        if self.parent_data:
            par = sum(len(person_data['children']) for person_data in self.selected_parent.values())
            print(par)
            self.par = str(par)

        else:
            self.par = "0"

        if self.visitor_data:
            vis = sum(int(entry.get('number_of_visitors', 0)) for entry in self.selected_visitor.values())
            self.vis = str(vis)

        else:
            self.vis = "0"

        if self.school_data:
            sch = sum(int(entry.get('number_of_students', 0)) for entry in self.selected_school.values())
            self.sch = str(sch)

            count = 0
            for i in self.selected_school:
                count += 1

            self.count = str(count)

        else:
            self.sch = "0"

        total = int(self.vis) + int(self.par) + int(self.sch)

        self.root.ids.tot.text = str(total)
        self.root.ids.visitor.text = str(self.vis)
        self.root.ids.school.text = str(self.count)
        self.root.ids.student.text = str(int(self.par) + int(self.sch))


    def month(self):
        current_time = datetime.now()
        month_name = current_time.strftime("%B")
        date_str = current_time.strftime("%Y-%m-%d")
        self.root.ids.date_label.text = date_str
        self.root.ids.month.text = month_name


    def start(self):
        try:
            self.root.ids.spine.active = True  # Access the screen first then the spinner
        except AttributeError:
            print("Error: Spinner or Screen not found. Check KV and IDs.")
        self.root.ids.spine.active = True

    def stop(self):
        self.root.ids.spine.active = False



    def show_date_picker(self):
        self.theme_cls.primary_palette = "Blue"
        date_dialog = MDDatePicker(min_year=2023, max_year=2031, primary_color="blue")
        date_dialog.bind(on_save=self.on_save, on_cancel=self.on_cancel)
        date_dialog.open()

    def on_save(self, instance, value, date_range):
        self.root.ids.date_label.text = f"Selected Date: {value}"
        self.selected_date = str(value)
        self.start()
        self.get_date()


    def on_cancel(self, instance, value):
        # Handle the cancel event (optional)
        self.root.ids.date_label.text = "Selection Canceled"
        self.root.ids.date_label.text_color =  (1, 0, 0, .5)

    def report_date_picker(self):
        self.theme_cls.primary_palette = "Orange"
        date_dialog = MDDatePicker(min_year=2025, max_year=2031, mode="range", primary_color="orange")
        date_dialog.bind(on_save=self.on_savu, on_cancel=self.on_cance)
        date_dialog.open()

        self.generate_excel("date")

    def on_savu(self, instance, value, date_range):
        self.generate_excel(value)

    def on_cance(self, instance, value):
        # Handle the cancel event (optional)
        pass

    def choice(self, name):
        print("choice")
        if name == "Parents":
            # screen_manager.current = "parent"
            screen_manager = self.root.ids.screen_manager
            screen_manager.current = "student"

        elif name == "Visitors":
            screen_manager = self.root.ids.screen_manager
            screen_manager.current = "visitor"

        elif name == "Schools":
            screen_manager = self.root.ids.screen_manager
            screen_manager.current = "school"

    def search(self):
        search_text = self.root.ids.search_field.text
        print(f"Searching for: {search_text}")

    def student_choice(self, name, pname, pnumber, cage):

        self.root.ids.parent_name.text = pname
        self.root.ids.contact.text = pnumber
        self.root.ids.child.text = name
        self.root.ids.age.text = cage + "  Years old"


        screen_manager = self.root.ids.screen_manager
        screen_manager.current = "parinfo"

    def visitor_choice(self, name, gname, total , phone):
        self.root.ids.sname.text = name
        self.root.ids.gui.text = gname
        self.root.ids.no.text = phone
        self.root.ids.total.text = total

        screen_manager = self.root.ids.screen_manager
        screen_manager.current = "visinfo"

    def school_choice(self, name, gname, total , phone):
        self.root.ids.wname.text = name
        self.root.ids.wgui.text = gname
        self.root.ids.wno.text = phone
        self.root.ids.wtotal.text = total

        screen_manager = self.root.ids.screen_manager
        screen_manager.current = "schoolinfo"

    @mainthread
    def get_date(self, ):
        year, smonth ,  day = self.selected_date.strip().split('-')

        year = str(year)
        smonth = int(smonth)
        day = str(int(day))


        months = [
            "January", "February", "March", "April", "May", "June",
            "July", "August", "September", "October", "November", "December"
        ]
        month_name = str(months[smonth - 1])


        data = Uz.selected_parent(Uz(), year, month_name, day)
        data2 =Uz.selected_guest(Uz(), year, month_name, day)
        data3 = Uz.selected_school(Uz(), year, month_name, day)

        if data:
            self.parent_data = data
        else:
            self.parent_data = ""

        if data2:
            self.visitor_data = data2
        else:
            self.visitor_data = ""

        if data3:
            self.school_data = data3
        else:
            self.school_data = ""

        self.display()

    def generate_excel(self, date):
        try:
            # Data for the report
            data = {
                "2024-12-01": {"total_schools": 3, "total_students": 80, "total_companies": 2, "total_visitors": 25},
                "2024-12-02": {"total_schools": 4, "total_students": 95, "total_companies": 1, "total_visitors": 10},
                "2024-12-03": {"total_schools": 5, "total_students": 120, "total_companies": 3, "total_visitors": 40},
            }

            # Totals
            totals = {
                "total_schools": sum(item["total_schools"] for item in data.values()),
                "total_students": sum(item["total_students"] for item in data.values()),
                "total_companies": sum(item["total_companies"] for item in data.values()),
                "total_visitors": sum(item["total_visitors"] for item in data.values()),
            }

            # Create an Excel workbook and sheet
            wb = Workbook()
            ws = wb.active
            ws.title = self.report

            # Header Row
            headers = ["Date", "Total Schools", "Total Students", "Total Companies", "Total Visitors"]
            ws.append(headers)

            # Add Daily Data
            for date, info in data.items():
                ws.append([
                    date,
                    info["total_schools"],
                    info["total_students"],
                    info["total_companies"],
                    info["total_visitors"],
                ])

            # Add Totals Row
            ws.append([
                "Totals",
                totals["total_schools"],
                totals["total_students"],
                totals["total_companies"],
                totals["total_visitors"]
            ])

            # Styling the Header
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            for col_num, col in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            # Enable and Configure Protection
            ws.protection.sheet = True
            ws.protection.set_password("securepassword")  # Replace with your desired password
            ws.protection.enable()

            # Save the file to external storage
            directory = "/storage/emulated/0/Download"
            if not os.path.exists(directory):
                os.makedirs(directory)
            filepath = os.path.join(directory, self.report + ".xlsx")
            wb.save(filepath)

            # Show a success message
            toast("Report Successful")
            self.filepath = filepath
            return filepath
        except Exception as e:
            toast("Error creating Report")
            return None


    def location(self):
        self.root.ids.path.text = self.filepath

    @mainthread
    def request_android_permissions(self):
        from android.permissions import request_permissions, Permission

        def callback(permissions, results):
            if all([res for res in results]):
                toast("callback. All permissions granted.")
            else:
                toast("callback. Some permissions refused.")

        request_permissions([Permission.WRITE_EXTERNAL_STORAGE, Permission.READ_EXTERNAL_STORAGE], callback)

    def register_caller(self, ):
        with open("parent.json", "w") as file:
            data = ["Uz.vistor(Uz())"]
            data_dump = json.dumps(data, indent=2)
            file.write(data_dump)
            file.close()


    def load(self, data_file_name):
        with open(data_file_name, "r") as file:
            initial_data = json.load(file)
        return initial_data


MainApp().run()